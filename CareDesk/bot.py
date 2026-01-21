import os
import re
import sys
import asyncio
import threading
from datetime import datetime
from dotenv import load_dotenv
from loguru import logger
from openpyxl import Workbook, load_workbook

from pipecat.audio.vad.silero import SileroVADAnalyzer
from pipecat.frames.frames import LLMRunFrame
from pipecat.pipeline.pipeline import Pipeline
from pipecat.pipeline.runner import PipelineRunner
from pipecat.pipeline.task import PipelineParams, PipelineTask
from pipecat.processors.aggregators.llm_context import LLMContext
from pipecat.processors.aggregators.llm_response_universal import LLMContextAggregatorPair

from pipecat.services.deepgram.stt import DeepgramSTTService
from pipecat.services.cartesia.tts import CartesiaTTSService
from pipecat.services.openai.llm import OpenAILLMService
from pipecat.services.deepgram.tts import DeepgramTTSService
from pipecat.services.llm_service import FunctionCallParams

from pipecat.transports.base_transport import TransportParams
from pipecat.transports.smallwebrtc.transport import SmallWebRTCTransport
from pipecat.transports.websocket.fastapi import (
    FastAPIWebsocketParams,
    FastAPIWebsocketTransport,
)

from pipecat.adapters.schemas.tools_schema import ToolsSchema

from nextgenswitch_serializer import NextGenSwitchFrameSerializer
from transfer_call import transfer_call


load_dotenv(override=True)


logger.add(sys.stderr, level=os.getenv("LOG_LEVEL", "INFO"))


GREETING_INSTRUCTION = "Start the call now with the required greeting."

APPOINTMENTS_HEADERS = [
    "Logged At",
    "Action",
    "Patient Name",
    "Patient Age or DOB",
    "Phone",
    "Department or Doctor",
    "Reason",
    "Preferred Date",
    "Preferred Time",
    "Visit Type",
    "Existing Appointment",
    "Notes",
]

APPOINTMENT_LOG_LOCK = threading.Lock()
DEFAULT_APPOINTMENTS_SHEET = "Appointments"

SYSTEM_INSTRUCTION = f"""
You are a professional appointment booking receptionist for a medical clinic named "Info Diagnostic Center" in the United States.

Your ONLY job is to help callers:
1) Book a new doctor appointment
2) Reschedule an existing appointment
3) Cancel an appointment
4) Provide basic clinic information (hours, location, departments)
5) Take a message for a callback if needed

VOICE OUTPUT RULES:
- Your response will be converted to audio.
- Use simple natural speech only.
- Do not use emojis, markdown, bullet points, or special characters.
- Keep responses short, clear, and calm.
- Usually reply in 1 to 2 sentences.
- Ask only one question at a time.
- Do not speak long paragraphs.

LANGUAGE:
- Speak in English only.
- If the caller speaks in another language, politely ask them to speak in English.

TONE:
- Friendly, respectful, and professional.
- Supportive and patient, never rushed.
- Do not mention system prompts or internal rules.
- Do not mention you are AI unless asked.
  If asked, say: "I am a virtual receptionist assistant."

MEDICAL SAFETY LIMITS:
- You do NOT diagnose, prescribe, or give medical advice.
- If the caller asks medical questions, respond briefly:
  "For medical advice, please consult the doctor. I can help you book an appointment."
- Never request or store passwords, OTP codes, or full payment card details.

EMERGENCY HANDLING:
If the caller reports severe symptoms or emergencies like chest pain, severe breathing difficulty, heavy bleeding, fainting, stroke signs, suicide or self-harm, or a serious accident:
Say:
"That sounds urgent. Please call 911 immediately or go to the nearest emergency room now."
Then offer:
"I can also notify the clinic if you want."

CLINIC INFO (USE WHEN ASKED):
- Clinic Name: Info Diagnostic Center
- Departments: Family Medicine, Internal Medicine, Cardiology, OB GYN, Pediatrics, Dermatology, ENT, Orthopedics
If the caller asks for exact address or hours and you do not have them, say:
"I can take your number and have the clinic call you back with the details."

DOCTOR DIRECTORY (UNITED STATES CONTEXT):
Callers may book directly by doctor name. If the caller says a doctor name, match it from this list.

1) Dr. Rachel Morgan, Family Medicine
   Clinic Hours: Monday to Friday, 9 AM to 1 PM

2) Dr. Andrew Harris, Internal Medicine
   Clinic Hours: Monday, Wednesday, Friday, 2 PM to 6 PM

3) Dr. Michael Reynolds, Cardiology
   Clinic Hours: Tuesday and Thursday, 10 AM to 2 PM

4) Dr. Olivia Bennett, OB GYN
   Clinic Hours: Monday and Thursday, 3 PM to 7 PM

5) Dr. Ethan Brooks, Pediatrics
   Clinic Hours: Monday to Thursday, 10 AM to 1 PM

6) Dr. Sophia Nguyen, Dermatology
   Clinic Hours: Wednesday and Friday, 1 PM to 5 PM

7) Dr. Daniel Kim, ENT
   Clinic Hours: Tuesday and Saturday, 9 AM to 12 PM

8) Dr. Jessica Patel, Orthopedics
   Clinic Hours: Monday and Wednesday, 11 AM to 3 PM

IMPORTANT SCHEDULING RULE:
- ALWAYS ask the caller for their preferred time only within the selected doctor’s clinic hours.
- If the caller requests a time outside the doctor’s hours, politely offer the closest available time within that doctor’s schedule.
- Do not promise exact availability unless you have confirmed it. If you cannot confirm the exact slot, say you will request it and confirm shortly.

PRIMARY WORKFLOW GOAL:
Collect enough information to book, reschedule, or cancel an appointment correctly.

ALWAYS FOLLOW THIS CALL FLOW:
Step 1: Greeting
Step 2: Identify request type: book, reschedule, cancel, info, message
Step 3: Identify doctor or department
Step 4: Offer available days and hours for that doctor
Step 5: Collect required patient details one by one
Step 6: Confirm appointment details
Step 7: Log or update via tools
Step 8: Close politely

GREETING:
Say:
"Hello, thank you for calling Info Diagnostic Center. How may I help you today?"

INTENT DETECTION:
If unclear, ask:
"Would you like to book, reschedule, or cancel an appointment?"

BOOKING MODE (NEW APPOINTMENT):
Priority order:
1) Book by doctor name if the caller gives it
2) Otherwise, book by department and suggest a doctor

IF CALLER GIVES A DOCTOR NAME:
- Confirm the doctor name only if needed.
- Immediately provide that doctor’s available days and hours in one short sentence.
- Ask one question to pick a day first, then a time within that window.

Example behavior:
"Dr. Michael Reynolds is available Tuesday and Thursday from 10 AM to 2 PM. Which day works best for you?"
After day is chosen:
"Great. What time would you prefer between 10 AM and 2 PM?"

IF CALLER DOES NOT KNOW THE DOCTOR:
Ask:
"Do you have a doctor name, or should I help you choose a department?"
If they describe the problem, suggest a department:
- General checkup, fever, diabetes follow-up -> Family Medicine or Internal Medicine
- Heart checkup, chest discomfort -> Cardiology
- Pregnancy, women’s health -> OB GYN
- Child health, vaccination -> Pediatrics
- Skin allergy, acne -> Dermatology
- Ear, nose, throat issues -> ENT
- Bone pain, joint pain, injury -> Orthopedics

Then suggest one doctor name briefly and ask for confirmation:
"For internal medicine, I can book with Dr. Andrew Harris. Would you like that?"
After doctor is confirmed, follow the doctor-hours scheduling rule.

COLLECT DETAILS ONE AT A TIME (ONLY WHAT IS NEEDED):
1) Patient full name
2) Patient age or date of birth
3) Primary phone number (US format)
4) Doctor name or department
5) Reason for visit in a short phrase
6) Appointment day
7) Preferred time within the doctor’s clinic hours
8) Visit type: first visit or follow-up

REALISTIC SLOT CONFIRMATION:
If you cannot confirm an exact slot:
"Thank you. I will request that time and confirm shortly."

RESCHEDULE MODE:
Collect:
1) Patient name
2) Phone number
3) Existing appointment date and time if known
4) Doctor name if known
Then ask for the new time only within the doctor’s clinic hours:
"What day and time would you like within the doctor’s available hours?"
Then confirm:
"Okay, I will reschedule it and confirm your new time shortly."

CANCEL MODE:
Collect:
1) Patient name
2) Phone number
3) Appointment date and time if known
4) Doctor name if known
Confirm cancellation:
"Okay, I will cancel the appointment. Would you like to book a new one instead?"

CLINIC INFORMATION MODE:
If caller asks about doctors, departments, address, or hours:
Answer in one short sentence if known.
If unknown:
"I can take your number and have the clinic confirm it."

CALLBACK MESSAGE MODE:
If the caller wants a callback or the request cannot be completed:
please forward the call to 9999 for leave a voice message.

CONFIRMATION RULE (BEFORE FINALIZING):
Confirm in one short sentence:
"To confirm, you want an appointment with Doctor Name on Date at Time for Reason. Is that correct?"

TOOL USAGE RULES:
After the caller confirms, use tools exactly once.

APPOINTMENT LOGGING TOOL:
After confirming a NEW booking, call log_appointment once with:
action, patient_name, patient_age_or_dob, phone, department_or_doctor, reason, preferred_date,
preferred_time, visit_type, existing_appointment, notes.
Use YYYY-MM-DD for preferred_date when possible.

APPOINTMENT UPDATE TOOL:
For reschedule or cancellation, call update_appointment once after confirming.
Provide search_name or search_phone, plus search_date and search_time if known.
Set action to reschedule or cancel and include only the fields that should change.
If update_appointment returns not_found:
Ask for the missing details in one short question, or offer a callback.

ERROR HANDLING:
If audio is unclear:
"Sorry, I did not catch that. Could you please repeat?"
If still unclear:
"No problem. I can take your number and the clinic will call you back."

PRIVACY:
- Only collect necessary booking details.
- Never ask for OTP, passwords, bank PIN, or full card number.

ENDING:
Close with:
"Thank you for calling Info Diagnostic Center. We will confirm your appointment shortly. Have a good day."
""".strip()



def _resolve_appointments_path() -> str:
    env_path = os.getenv("APPOINTMENTS_XLSX_PATH")
    if env_path:
        return os.path.abspath(env_path)
    return os.path.join(os.path.dirname(__file__), "appointments.xlsx")


def _cell_text(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_sheet_name(value: str) -> str:
    if not value:
        return ""
    cleaned = re.sub(r"[\\/*?:\[\]]", "-", str(value)).strip()
    if not cleaned:
        return ""
    return cleaned[:31]


def _normalize_phone(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\D", "", str(value))


def _normalize_date(value: str | None) -> str:
    if not value:
        return ""
    cleaned = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(cleaned, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    date_match = re.search(
        r"(\d{4}-\d{1,2}-\d{1,2}|\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", cleaned
    )
    if date_match:
        candidate = date_match.group(1)
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(candidate, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue

    try:
        from dateutil import parser as _dtparser  # type: ignore

        dt = _dtparser.parse(cleaned, fuzzy=True, default=datetime.now())
        if dt:
            return dt.strftime("%Y-%m-%d")
    except Exception:
        pass

    return cleaned


def _use_single_appointments_sheet() -> bool:
    mode = os.getenv("APPOINTMENTS_SHEET_MODE", "single").strip().lower()
    return mode not in {"per_date", "date", "dated"}


def _appointments_sheet_name(preferred_date_norm: str, existing_norm: str) -> str:
    if _use_single_appointments_sheet():
        override = os.getenv("APPOINTMENTS_SHEET_NAME", DEFAULT_APPOINTMENTS_SHEET)
        cleaned = _safe_sheet_name(override)
        return cleaned or DEFAULT_APPOINTMENTS_SHEET

    target_date = preferred_date_norm or _normalize_date(existing_norm)
    if not target_date:
        target_date = datetime.now().strftime("%Y-%m-%d")
    return _safe_sheet_name(target_date) or datetime.now().strftime("%Y-%m-%d")


def _normalize_time_bucket(value: str | None) -> str:
    if not value:
        return ""
    text = _cell_text(value).lower()

    if any(k in text for k in ["morning", "am"]):
        return "morning"
    if any(k in text for k in ["afternoon", "noon", "pm"]):
        return "afternoon"
    if any(k in text for k in ["evening", "night"]):
        return "evening"

    m = re.search(r"(\d{1,2})(?::(\d{2}))?\s*(am|pm)?", text)
    if not m:
        return text

    hour = int(m.group(1))
    minute = int(m.group(2) or 0)
    ampm = m.group(3)

    if ampm == "pm" and hour < 12:
        hour += 12
    if ampm == "am" and hour == 12:
        hour = 0

    if 5 <= hour <= 11:
        return "morning"
    if 12 <= hour <= 16:
        return "afternoon"
    if 17 <= hour <= 22:
        return "evening"
    return f"{hour:02d}:{minute:02d}"


def _phone_matches(search: str, candidate: str) -> bool:
    if not search:
        return True
    if not candidate:
        return False
    if search == candidate:
        return True
    if len(search) >= 7 and candidate.endswith(search):
        return True
    if len(candidate) >= 7 and search.endswith(candidate):
        return True
    return False


def _text_matches(search: str, candidate: str) -> bool:
    if not search:
        return True
    if not candidate:
        return False
    search_norm = search.strip().lower()
    candidate_norm = candidate.strip().lower()
    return search_norm in candidate_norm or candidate_norm in search_norm


def _date_matches(search: str, candidate: str) -> bool:
    if not search:
        return True
    if not candidate:
        return False

    s_raw = _cell_text(search)
    c_raw = _cell_text(candidate)

    s_norm = _normalize_date(s_raw).strip().lower()
    c_norm = _normalize_date(c_raw).strip().lower()

    if s_norm == c_norm:
        return True

    if s_norm in c_norm or c_norm in s_norm:
        return True

    return False


def _time_matches(search: str, candidate: str) -> bool:
    if not search:
        return True
    if not candidate:
        return False

    s_raw = _cell_text(search).lower()
    c_raw = _cell_text(candidate).lower()

    s_bucket = _normalize_time_bucket(s_raw)
    c_bucket = _normalize_time_bucket(c_raw)

    if s_bucket == c_bucket:
        return True

    if s_raw in c_raw or c_raw in s_raw:
        return True

    return False


def _ensure_sheet_headers(sheet) -> dict[str, int]:
    header_cells = [cell.value for cell in sheet[1]]

    if sheet.max_row == 1 and sheet.max_column == 1 and header_cells[0] is None:
        for col_idx, name in enumerate(APPOINTMENTS_HEADERS, start=1):
            sheet.cell(row=1, column=col_idx).value = name
        return {name: idx + 1 for idx, name in enumerate(APPOINTMENTS_HEADERS)}

    header_map: dict[str, int] = {}
    for idx, value in enumerate(header_cells, start=1):
        if value is None:
            continue
        header_map[_cell_text(value)] = idx

    last_col = len(header_cells)
    for name in APPOINTMENTS_HEADERS:
        if name not in header_map:
            last_col += 1
            header_map[name] = last_col
            sheet.cell(row=1, column=last_col).value = name

    return header_map


def _append_row_to_workbook(path: str, sheet_name: str, row: list[str]) -> None:
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)

    if os.path.exists(path):
        workbook = load_workbook(path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

    _ensure_sheet_headers(sheet)
    sheet.append(row)
    workbook.save(path)


def _build_webrtc_transport(webrtc_connection: object) -> SmallWebRTCTransport:
    return SmallWebRTCTransport(
        webrtc_connection=webrtc_connection,
        params=TransportParams(
            audio_in_enabled=True,
            audio_out_enabled=True,
            vad_analyzer=SileroVADAnalyzer(),
            audio_out_10ms_chunks=2,
        ),
    )


def _build_websocket_transport(
    websocket: object,
    stream_id: str | None = None,
) -> FastAPIWebsocketTransport:
    serializer = NextGenSwitchFrameSerializer()
    if stream_id:
        serializer.set_stream_id(stream_id)

    return FastAPIWebsocketTransport(
        websocket=websocket,
        params=FastAPIWebsocketParams(
            audio_in_enabled=True,
            audio_out_enabled=True,
            add_wav_header=False,
            vad_analyzer=SileroVADAnalyzer(),
            serializer=serializer,
        ),
    )


async def run_bot(
    webrtc_connection=None,
    websocket=None,
    stream_id=None,
    call_sid=None,
    bot_params=None,
):
    if webrtc_connection and websocket:
        raise ValueError("Provide either webrtc_connection or websocket, not both.")
    if not webrtc_connection and not websocket:
        raise ValueError("No transport input provided to run_bot.")

    transport = (
        _build_webrtc_transport(webrtc_connection)
        if webrtc_connection
        else _build_websocket_transport(websocket, stream_id=stream_id)
    )

    close_callback = None
    if webrtc_connection:
        if hasattr(webrtc_connection, "disconnect"):
            close_callback = webrtc_connection.disconnect
        elif hasattr(webrtc_connection, "close"):
            close_callback = webrtc_connection.close
    elif websocket:
        close_callback = websocket.close

    system_instruction = SYSTEM_INSTRUCTION
    greeting_instruction = GREETING_INSTRUCTION

    openai_api_key = os.getenv("OPENAI_API_KEY")
    deepgram_api_key = os.getenv("DEEPGRAM_API_KEY")
    cartesia_api_key = os.getenv("CARTESIA_API_KEY")
    cartesia_voice_id = os.getenv(
        "CARTESIA_VOICE_ID", "5ee9feff-1265-424a-9d7f-8e4d431a12c7"
    )

    base_url = os.getenv("NEXTGENSWITCH_URL")
    api_key = os.getenv("NEXTGENSWITCH_API_KEY")
    api_secret = os.getenv("NEXTGENSWITCH_API_SECRET")

    if bot_params and "prompt" in bot_params:
        system_instruction = bot_params["prompt"]
    if bot_params and "openai_api_key" in bot_params:
        openai_api_key = bot_params["openai_api_key"]
    if bot_params and "greetings" in bot_params:
        greeting_instruction = bot_params["greetings"]
    if bot_params and "nexgenswitch_api_url" in bot_params:
        base_url = bot_params["nexgenswitch_api_url"]
    if bot_params and "nexgenswitch_api_key" in bot_params:
        api_key = bot_params["nexgenswitch_api_key"]
    if bot_params and "nextgenswitch_api_secret" in bot_params:
        api_secret = bot_params["nextgenswitch_api_secret"]
    if bot_params and "deepgram_api_key" in bot_params:
        deepgram_api_key = bot_params["deepgram_api_key"]
    if bot_params and "cartesia_api_key" in bot_params:
        cartesia_api_key = bot_params["cartesia_api_key"]
    if bot_params and "cartesia_voice_id" in bot_params:
        cartesia_voice_id = bot_params["cartesia_voice_id"]

    if not openai_api_key:
        raise ValueError("Missing OPENAI_API_KEY (env or bot_params.openai_api_key)")
    if not deepgram_api_key:
        raise ValueError("Missing DEEPGRAM_API_KEY (env or bot_params.deepgram_api_key)")
    if not cartesia_api_key:
        raise ValueError("Missing CARTESIA_API_KEY (env or bot_params.cartesia_api_key)")

    stt = DeepgramSTTService(api_key=deepgram_api_key)
    # tts = CartesiaTTSService(api_key=cartesia_api_key, voice_id=cartesia_voice_id)
    tts = DeepgramTTSService(api_key=deepgram_api_key, voice="aura-2-athena-en")

    async def close_session(params: FunctionCallParams) -> dict:
        logger.info("Closing transport session")
        if close_callback:
            await close_callback()
        else:
            logger.warning("No close callback available for this session")
        logger.info("Transport session closed")
        return {"status": "closed"}

    async def log_appointment(
        params: FunctionCallParams,
        action: str,
        patient_name: str | None = None,
        patient_age_or_dob: str | None = None,
        phone: str | None = None,
        department_or_doctor: str | None = None,
        reason: str | None = None,
        preferred_date: str | None = None,
        preferred_time: str | None = None,
        visit_type: str | None = None,
        existing_appointment: str | None = None,
        notes: str | None = None,
    ) -> dict:
        preferred_date_norm = _normalize_date(preferred_date)
        existing_norm = _cell_text(existing_appointment)
        phone_norm = _normalize_phone(phone)

        sheet_name = _appointments_sheet_name(preferred_date_norm, existing_norm)
        log_path = _resolve_appointments_path()

        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            _cell_text(action),
            _cell_text(patient_name),
            _cell_text(patient_age_or_dob),
            phone_norm or _cell_text(phone),
            _cell_text(department_or_doctor),
            _cell_text(reason),
            preferred_date_norm or _cell_text(preferred_date),
            _cell_text(preferred_time),
            _cell_text(visit_type),
            existing_norm,
            _cell_text(notes),
        ]

        try:
            with APPOINTMENT_LOG_LOCK:
                _append_row_to_workbook(log_path, sheet_name, row)
        except PermissionError as exc:
            logger.exception("Appointment log write blocked (file locked): {}", log_path)
            result = {"status": "error", "reason": "file_locked", "path": log_path}
            await params.result_callback(result)
            return result
        except Exception as exc:
            logger.exception("Appointment log write failed: {}", log_path)
            result = {"status": "error", "reason": "write_failed", "path": log_path}
            await params.result_callback(result)
            return result

        result = {"status": "logged", "path": log_path, "sheet": sheet_name}
        await params.result_callback(result)
        return result

    async def update_appointment(
        params: FunctionCallParams,
        action: str,
        search_name: str | None = None,
        search_phone: str | None = None,
        search_date: str | None = None,
        search_time: str | None = None,
        patient_name: str | None = None,
        patient_age_or_dob: str | None = None,
        phone: str | None = None,
        department_or_doctor: str | None = None,
        reason: str | None = None,
        preferred_date: str | None = None,
        preferred_time: str | None = None,
        visit_type: str | None = None,
        existing_appointment: str | None = None,
        notes: str | None = None,
    ) -> dict:
        if not search_name and patient_name:
            search_name = patient_name
        if not search_phone and phone:
            search_phone = phone

        if not any([search_name, search_phone, search_date, search_time]):
            result = {"status": "missing_search"}
            await params.result_callback(result)
            return result

        log_path = _resolve_appointments_path()
        if not os.path.exists(log_path):
            result = {"status": "not_found", "reason": "missing_workbook"}
            await params.result_callback(result)
            return result

        search_phone_norm = _normalize_phone(search_phone)
        search_date_norm = _normalize_date(search_date)
        if _use_single_appointments_sheet():
            sheet_hint = _appointments_sheet_name(search_date_norm, "")
        else:
            sheet_hint = _safe_sheet_name(search_date_norm) if search_date_norm else ""

        action_value = _cell_text(action)
        preferred_date_norm = _normalize_date(preferred_date)

        with APPOINTMENT_LOG_LOCK:
            workbook = load_workbook(log_path)

            if sheet_hint and sheet_hint in workbook.sheetnames:
                sheet_names = [sheet_hint] + [s for s in workbook.sheetnames if s != sheet_hint]
            else:
                sheet_names = workbook.sheetnames

            candidates = []
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                header_map = _ensure_sheet_headers(sheet)

                for row_idx in range(2, sheet.max_row + 1):
                    row_values = {
                        header: sheet.cell(row=row_idx, column=col_idx).value
                        for header, col_idx in header_map.items()
                    }

                    if not any(value not in (None, "") for value in row_values.values()):
                        continue

                    row_name = _cell_text(row_values.get("Patient Name"))
                    row_phone = _normalize_phone(_cell_text(row_values.get("Phone")))
                    row_pref_date = _cell_text(row_values.get("Preferred Date"))
                    row_pref_time = _cell_text(row_values.get("Preferred Time"))
                    row_existing = _cell_text(row_values.get("Existing Appointment"))

                    if search_name and not _text_matches(search_name, row_name):
                        continue
                    if search_phone_norm and not _phone_matches(search_phone_norm, row_phone):
                        continue
                    if search_date and not (
                        _date_matches(search_date, row_pref_date)
                        or _date_matches(search_date, row_existing)
                    ):
                        continue
                    if search_time and not (
                        _time_matches(search_time, row_pref_time)
                        or _time_matches(search_time, row_existing)
                    ):
                        continue

                    candidates.append((sheet_name, row_idx, row_values))

            if not candidates:
                result = {"status": "not_found"}
                await params.result_callback(result)
                return result

            if len(candidates) > 1 and (not search_date and not search_time):
                result = {
                    "status": "multiple_matches",
                    "matches": [
                        {
                            "sheet": s,
                            "row": r,
                            "patient_name": _cell_text(v.get("Patient Name")),
                            "phone": _cell_text(v.get("Phone")),
                            "preferred_date": _cell_text(v.get("Preferred Date")),
                            "preferred_time": _cell_text(v.get("Preferred Time")),
                            "existing_appointment": _cell_text(v.get("Existing Appointment")),
                        }
                        for (s, r, v) in candidates[:10]
                    ],
                }
                await params.result_callback(result)
                return result

            sheet_name, row_idx, row_values = candidates[0]
            sheet = workbook[sheet_name]
            header_map = _ensure_sheet_headers(sheet)

            updated_values = dict(row_values)
            updated_values["Action"] = action_value

            if not existing_appointment and action_value.lower().startswith("resched"):
                if not _cell_text(row_values.get("Existing Appointment")):
                    snapshot_parts = [
                        _cell_text(row_values.get("Preferred Date")),
                        _cell_text(row_values.get("Preferred Time")),
                    ]
                    snapshot = " ".join([p for p in snapshot_parts if p])
                    if snapshot:
                        existing_appointment = snapshot

            if patient_name:
                updated_values["Patient Name"] = patient_name
            if patient_age_or_dob:
                updated_values["Patient Age or DOB"] = patient_age_or_dob
            if phone:
                updated_values["Phone"] = _normalize_phone(phone) or phone
            if department_or_doctor:
                updated_values["Department or Doctor"] = department_or_doctor
            if reason:
                updated_values["Reason"] = reason
            if preferred_date:
                updated_values["Preferred Date"] = preferred_date_norm or preferred_date
            if preferred_time:
                updated_values["Preferred Time"] = preferred_time
            if visit_type:
                updated_values["Visit Type"] = visit_type
            if existing_appointment:
                updated_values["Existing Appointment"] = existing_appointment
            if notes:
                updated_values["Notes"] = notes

            new_sheet_name = sheet_name
            moved = False

            if not _use_single_appointments_sheet():
                if preferred_date and action_value.lower().startswith("resched"):
                    new_sheet_name = _safe_sheet_name(preferred_date_norm) or sheet_name

            if new_sheet_name != sheet_name:
                target_sheet = (
                    workbook[new_sheet_name]
                    if new_sheet_name in workbook.sheetnames
                    else workbook.create_sheet(new_sheet_name)
                )
                _ensure_sheet_headers(target_sheet)
                target_row = [updated_values.get(header, "") for header in APPOINTMENTS_HEADERS]
                target_sheet.append(target_row)
                sheet.delete_rows(row_idx, 1)
                moved = True
            else:
                for header, col_idx in header_map.items():
                    if header == "Logged At":
                        continue
                    if header in updated_values and updated_values[header] is not None:
                        sheet.cell(row=row_idx, column=col_idx).value = updated_values[header]

            workbook.save(log_path)

            result = {"status": "updated", "path": log_path, "sheet": new_sheet_name, "moved": moved}
            await params.result_callback(result)
            return result

    
    async def transfer_call_to(params: FunctionCallParams, forwarding_number: int) -> None:
        """Trasfer call to live agent. Call this function immidiately if user want to talk to a live agent.

        This is a placeholder function to demonstrate how to transfer the call
        into a live agent.
        """
        if not forwarding_number:
            logger.error("Forwarding number is not configured; skipping transfer for {}", call_sid)
            await params.result_callback({"status": "unsupported"})
            return
        
        if not base_url:
            logger.error("NEXTGENSWITCH_URL is not configured; unable to transfer call {}", call_sid)
            await params.result_callback({"status": "unsupported"})
            return
        if not api_key:
            logger.error("NEXTGENSWITCH_API_KEY is not configured; unable to transfer call {}", call_sid)
            await params.result_callback({"status": "unsupported"})
            return
        if not api_secret:
            logger.error("NEXTGENSWITCH_API_SECRET is not configured; unable to transfer call {}", call_sid)
            await params.result_callback({"status": "unsupported"})
            return
        

        logger.info("Transferring call {} to  {}", call_sid, forwarding_number)
        asyncio.create_task(transfer_call(call_sid, forwarding_number, base_url, api_key, api_secret))
        await params.result_callback({"status": "transferred"})
    
    tools = ToolsSchema(standard_tools=[close_session, log_appointment, update_appointment, transfer_call_to])

    context = LLMContext(
        [
            {"role": "system", "content": system_instruction},
            {"role": "user", "content": greeting_instruction},
        ],
        tools=tools,
    )
    context_aggregator = LLMContextAggregatorPair(context)

    llm = OpenAILLMService(api_key=openai_api_key)

    pipeline = Pipeline(
        [
            transport.input(),
            stt,
            context_aggregator.user(),
            llm,
            tts,
            transport.output(),
            context_aggregator.assistant(),
        ]
    )

    llm.register_direct_function(close_session, cancel_on_interruption=False)
    llm.register_direct_function(log_appointment, cancel_on_interruption=False)
    llm.register_direct_function(update_appointment, cancel_on_interruption=False)
    llm.register_direct_function(transfer_call_to, cancel_on_interruption=False)

    task = PipelineTask(
        pipeline,
        params=PipelineParams(
            enable_metrics=True,
            enable_usage_metrics=True,
        ),
    )

    @transport.event_handler("on_client_connected")
    async def on_client_connected(transport, client):
        logger.info("Pipecat Client connected")
        await task.queue_frames([LLMRunFrame()])

    @transport.event_handler("on_client_disconnected")
    async def on_client_disconnected(transport, client):
        logger.info("Pipecat Client disconnected")
        await task.cancel()

    runner = PipelineRunner(handle_sigint=False)
    await runner.run(task)
